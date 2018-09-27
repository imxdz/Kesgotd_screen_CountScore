# coding:utf-8
from openpyxl import load_workbook
import openpyxl


def copy_excel(excelpath1, excelpath2):#excelpath1就是要复制的excel地址，excelpath2就是复制到的excel地址
    '''复制excek，把excelpath1数据复制到excelpath2'''
    wb2 = openpyxl.Workbook()#生成一个实例，可以注释掉
    wb2.save(excelpath2)#生成一个空excel文件，可以注释掉
    # 读取数据
    wb1 = openpyxl.load_workbook(excelpath1)#读取要复制的excel
    wb2 = openpyxl.load_workbook(excelpath2)#读取复制到的excel
    sheets1 = wb1.sheetnames#获取要复制excel下所有sheet
    sheets2 = wb2.sheetnames#获取空excel下的sheet
    sheet1 = wb1[sheets1[0]]#获取第一个sheet
    sheet2 = wb2[sheets2[0]]#获取第一个sheet
    max_row = sheet1.max_row         # 要复制的excel的最大行数
    max_column = sheet1.max_column   # 要复制的excel的最大列数

    for m in list(range(1,max_row+1)):#从1开始循环到最大行数
        for n in list(range(97,97+max_column)):   # 循环a、b、c.....
            n = chr(n)                            # ASCII字符，chr(97)='a'
            i ='%s%d'% (n, m)                     # 单元格编号eg:A1、B1
            cell1 = sheet1[i].value               # 获取要复制的excel单元格数据
            sheet2[i].value = cell1               # 赋值到复制到的excel单元格

    wb2.save(excelpath2)                 # 保存数据
    wb1.close()                          # 关闭excel
    wb2.close()

class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename):
        self.wb = openpyxl.Workbook()#生成一个实例
        self.wb.save(filename)#生成一个空excel文件
        self.filename = filename
        self.wb = load_workbook(self.filename)#读取excel
        self.ws = self.wb.active  # 激活sheet

    def writee(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n).value = value#给excel的某行某列赋值
        self.wb.save(self.filename)#保存文件
        self.wb.close()#关闭文件
#总结：两个方法，一个是copy excel的所有值到另外一个excel，包括第一行的列名
if __name__ == "__main__":
    # copy_excel("api.xlsx", "test.xlsx")
    wt = Write_excel("2.xlsx")
    wt.writee(5, 7, "interface")
    wt.writee(5, 8, "interface")
    print(1)



